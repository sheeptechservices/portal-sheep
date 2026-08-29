"""
Import cedentes from Excel into the DUX admin API.

Usage:
    python scripts/import-cedentes.py <API_URL> <ADMIN_PASSWORD> [EXCEL_PATH]

Example (local dev):
    python scripts/import-cedentes.py http://localhost:3000/api/admin-data SUA_SENHA

Example (production):
    python scripts/import-cedentes.py https://dux-forms-solicitacoes.vercel.app/api/admin-data SUA_SENHA

pip install openpyxl requests
"""

import sys
import re
import json
import math
from datetime import datetime, date

try:
    import openpyxl
    import requests
except ImportError:
    print("Instale as dependências: pip install openpyxl requests")
    sys.exit(1)

API_URL  = sys.argv[1] if len(sys.argv) > 1 else "http://localhost:3000/api/admin-data"
PASSWORD = sys.argv[2] if len(sys.argv) > 2 else ""
EXCEL    = sys.argv[3] if len(sys.argv) > 3 else r"C:\Users\gui-z\Downloads\Pasta1.xlsx"
BATCH    = 50  # records per API call

# ── Helpers ───────────────────────────────────────────────────────────────────

EMOJI_RE = re.compile(
    r'[\U0001F300-\U0001F9FF'
    r'\U00002700-\U000027BF'
    r'\U0000FE00-\U0000FE0F'
    r'\U00002000-\U000021FF'
    r'‍️⃣]+',
    re.UNICODE
)

def clean(v):
    """Strip emojis and leading/trailing whitespace. Return None if empty."""
    if v is None:
        return None
    if isinstance(v, float) and math.isnan(v):
        return None
    s = str(v).strip()
    s = EMOJI_RE.sub('', s).strip()
    return s or None

def to_bool(v):
    if v is None:
        return 0
    return 1 if str(v).strip().upper() in ('TRUE', '1', 'SIM', 'S', 'YES') else 0

def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)) and not math.isnan(v):
        return float(v)
    s = str(v).replace(',', '.').strip()
    try:
        return float(s)
    except ValueError:
        return None

def to_date(v):
    if v is None:
        return datetime.now().isoformat()
    if isinstance(v, (datetime,)):
        return v.isoformat()
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day).isoformat()
    return str(v)

# Known flags mapping (strip emoji variants)
FLAGS_MAP = {
    'regular': 'Regular',
    'com pendências': 'Com pendências',
    'com pendencias': 'Com pendências',
    'inadimplente': 'Inadimplente',
    'sem operação': None,   # not in spec → null
    'sem operacao': None,
}

STATUS_MAP = {
    'ativo': 'Ativo',
    'parado': 'Parado',
    'banido': 'Banido',
}

# ── Read Excel ─────────────────────────────────────────────────────────────────

print(f"Lendo: {EXCEL}")
wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb.active

headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
print(f"Colunas encontradas ({len(headers)}): {headers[:10]}...")

# Column index map
col = {h: i for i, h in enumerate(headers)}

def get(row, name, alt=None):
    """Get cell value by header name."""
    idx = col.get(name, col.get(alt) if alt else None)
    if idx is None:
        return None
    return row[idx].value

rows = list(ws.iter_rows(min_row=2, values_only=False))
print(f"Total de linhas: {len(rows)}")

cedentes = []
for row in rows:
    nome = clean(get(row, 'NOME'))
    if not nome:
        continue  # skip empty rows

    flags_raw = clean(get(row, 'FLAGS'))
    flags_key = (flags_raw or '').lower()
    flags = FLAGS_MAP.get(flags_key, flags_raw)  # passthrough if not in map

    status_raw = clean(get(row, 'STATUS'))
    status_key = (status_raw or '').lower()
    status = STATUS_MAP.get(status_key, status_raw or 'Ativo')

    c = {
        'nome':               nome,
        'cnpj_cpf':           clean(get(row, 'CNPJ/CPF')),
        'razao_social':       clean(get(row, 'RAZÃO SOCIAL', 'RAZAO SOCIAL')),
        'status':             status,
        'flags':              flags,
        'origem':             clean(get(row, 'ORIGEM')),
        'segmento':           clean(get(row, 'SEGMENTO')),
        'sub_segmento':       clean(get(row, 'SUB-SEGMENTO')),
        'origem_comercial':   clean(get(row, 'ORIGEM COMERCIAL')),
        'canal_aquisicao':    clean(get(row, 'CANAL DE AQUISIÇÃO', 'CANAL DE AQUISICAO')),
        'parceiro':           to_bool(get(row, 'PARCEIRO?')),
        'cidade_estado':      clean(get(row, 'CIDADE/ESTADO')),
        'natureza_juridica':  clean(get(row, 'NATUREZA JURÍDICA', 'NATUREZA JURIDICA')),
        'valores_em_aberto':  None,  # coming soon
        'limite_operacao':    to_float(get(row, 'LIMITE DE OPERAÇÃO PARA CLIENTE', 'LIMITE DE OPERACAO PARA CLIENTE')),
        'rating':             to_float(get(row, 'RATING DO CLIENTE (0-10)')),
        'obs':                clean(get(row, 'OBS')),
        'email':              clean(get(row, 'E-MAIL')),
        'endereco_pj':        clean(get(row, 'ENDEREÇO PJ', 'ENDERECO PJ')),
        'nome_responsavel':   clean(get(row, 'NOME RESPONSÁVEL', 'NOME RESPONSAVEL')),
        'email_responsavel':  clean(get(row, 'EMAIL RESPONSÁVEL', 'EMAIL RESPONSAVEL')),
        'endereco_responsavel': clean(get(row, 'ENDEREÇO RESPONSÁVEL', 'ENDERECO RESPONSAVEL')),
        'cpf_responsavel':    clean(get(row, 'CPF RESPONSÁVEL', 'CPF RESPONSAVEL')),
        'possui_escrow':      to_bool(get(row, 'POSSUI ESCROW?', 'POSSUI CONTA ESCROW?')),
        'wpp_contato':        clean(get(row, 'WPP CONTATO')),
        'conta_escrow':       clean(get(row, 'Nº DA CONTA ESCROW')),
        'criado_em':          to_date(get(row, 'DATA CADASTRO')),
    }
    cedentes.append(c)

print(f"Cedentes válidos para importar: {len(cedentes)}")

# ── Import in batches ──────────────────────────────────────────────────────────

if not PASSWORD:
    print("\nATENÇÃO: Nenhuma senha informada. Passe a senha admin como 2º argumento.")
    sys.exit(1)

total_ok = 0
total_batches = math.ceil(len(cedentes) / BATCH)

for i in range(0, len(cedentes), BATCH):
    batch = cedentes[i:i + BATCH]
    batch_num = i // BATCH + 1
    print(f"Enviando lote {batch_num}/{total_batches} ({len(batch)} registros)...", end=' ', flush=True)
    try:
        res = requests.post(
            API_URL,
            headers={'Content-Type': 'application/json', 'x-admin-password': PASSWORD},
            json={'action': 'import_cedentes', 'cedentes': batch},
            timeout=60,
        )
        data = res.json()
        if data.get('ok'):
            total_ok += data.get('count', len(batch))
            print(f"OK ({data.get('count', '?')} inseridos)")
        else:
            print(f"ERRO: {data}")
    except Exception as e:
        print(f"FALHOU: {e}")

print(f"\nConcluído. Total inserido: {total_ok}/{len(cedentes)}")
